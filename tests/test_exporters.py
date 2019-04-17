import unittest
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook
from scrapy import Field, Item

from scrapy_xlsx import XlsxItemExporter


class XlsxItemExporterTest(unittest.TestCase):

    def setUp(self):
        class TestItem(Item):
            name = Field()
            age = Field()

        self.item = TestItem(name='John', age='42')
        self.output = NamedTemporaryFile()
        self.exporter = self._get_exporter()

    def tearDown(self):
        self.output.close()

    def _get_exporter(self, **kwargs):
        return XlsxItemExporter(self.output, **kwargs)

    def _check_basic_workflow(self, item):
        self.exporter.start_exporting()
        self.exporter.export_item(item)
        self.exporter.finish_exporting()

    def test_export_scrapy_item(self):
        self._check_basic_workflow(self.item)

    def test_export_dict_item(self):
        self._check_basic_workflow(dict(self.item))

    def test_fields_to_export(self):
        ie = self._get_exporter(fields_to_export=['name'])
        self.assertEqual(list(ie._get_serialized_fields(self.item)), [
                         ('name', 'John')])

    def test_serialize_field(self):
        res = self.exporter.serialize_field(
            self.item.fields['name'], 'name', self.item['name'])
        self.assertEqual(res, 'John')

        res = self.exporter.serialize_field(
            self.item.fields['age'], 'age', self.item['age'])
        self.assertEqual(res, '42')

    def test_serialize_field_join_multivalued(self):
        res = self.exporter.serialize_field({}, 'name', ['a', 'b', 'c'])
        self.assertEqual(res, 'a,b,c')

        exporter = self._get_exporter(join_multivalued=None)
        res = exporter.serialize_field({}, 'name', ['a', 'b', 'c'])
        self.assertEqual(res, "['a', 'b', 'c']")

        exporter = self._get_exporter(join_multivalued='|')
        res = exporter.serialize_field({}, 'name', ['a', 'b', 'c'])
        self.assertEqual(res, 'a|b|c')

        exporter = self._get_exporter(join_multivalued='|')
        res = exporter.serialize_field({}, 'name', ['a', 1])
        self.assertEqual(res, "['a', 1]")

    def test_field_custom_serializer(self):
        def custom_serializer(value):
            return value.lower()

        class CustomItem(Item):
            name = Field(serializer=custom_serializer)
            age = Field()

        item = CustomItem(name='John', age='42')
        self.assertEqual(self.exporter.serialize_field(
            item.fields['name'], 'name', item['name']), 'john')
        self.assertEqual(self.exporter.serialize_field(
            item.fields['age'], 'age', item['age']), '42')

    def test_output_content(self):
        exporter = self._get_exporter(fields_to_export=['age', 'name'])
        exporter.start_exporting()
        exporter.export_item(self.item)
        exporter.finish_exporting()

        workbook = load_workbook(self.output)
        sheet = workbook.active

        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet.max_column, 2)

        expected_rows = [
            ['age', 'name'],
            ['42', 'John'],
        ]
        for row, row_values in enumerate(sheet.iter_rows()):
            for column, cell in enumerate(row_values):
                self.assertEqual(cell.value, expected_rows[row][column])

    def test_output_content_no_header_row(self):
        exporter = self._get_exporter(
            include_header_row=False, fields_to_export=['age', 'name'])
        exporter.start_exporting()
        exporter.export_item(self.item)
        exporter.finish_exporting()

        workbook = load_workbook(self.output)
        sheet = workbook.active

        self.assertEqual(sheet.max_row, 1)
        self.assertEqual(sheet.max_column, 2)

        expected_rows = [
            ['42', 'John'],
        ]
        for row, row_values in enumerate(sheet.iter_rows()):
            for column, cell in enumerate(row_values):
                self.assertEqual(cell.value, expected_rows[row][column])

    def test_output_content_default_value(self):
        exporter = self._get_exporter(
            fields_to_export=['age', 'name'], default_value='-')
        exporter.start_exporting()
        exporter.export_item({'name': 'John'})
        exporter.finish_exporting()

        workbook = load_workbook(self.output)
        sheet = workbook.active

        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet.max_column, 2)

        expected_rows = [
            ['age', 'name'],
            ['-', 'John'],
        ]
        for row, row_values in enumerate(sheet.iter_rows()):
            for column, cell in enumerate(row_values):
                self.assertEqual(cell.value, expected_rows[row][column])
